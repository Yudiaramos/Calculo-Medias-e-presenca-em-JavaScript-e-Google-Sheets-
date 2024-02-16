import openpyxl #biblioteca que permite usra arquivos tipo xlsx em vez de csv

#definimos 3 funções que são as possibilidades das notas

def calcular_media(p1, p2, p3):
    return (float(p1) + float(p2) + float(p3)) / 3

def calcular_naf(media):
    return 100 - media

def reprovado_por_falta(total_faltas, total_aulas):
    return total_faltas > 0.25 * total_aulas

def calcular_situacao_aluno(linha):
    #sheet -> pega os valores das colunas e linhas especificadas
    p1 = sheet.cell(row=linha, column=4).value
    p2 = sheet.cell(row=linha, column=5).value
    p3 = sheet.cell(row=linha, column=6).value
    total_faltas = sheet.cell(row=linha, column=3).value
    total_aulas = 60  

    media = calcular_media(p1, p2, p3)

    if reprovado_por_falta(total_faltas, total_aulas):
        sheet.cell(row=linha, column=7).value = 'Reprovado por Falta'
    else:
        if media < 50:
            sheet.cell(row=linha, column=7).value = 'Reprovado por Nota'
        elif 50 <= media < 70:
            sheet.cell(row=linha, column=7).value = 'Exame Final'
            # caulo a Nota para Aprovação Final (naf)
            naf = calcular_naf(media)
            sheet.cell(row=linha, column=8).value = naf
        else:
            sheet.cell(row=linha, column=7).value = 'Aprovado'
            sheet.cell(row=linha, column=8).value = 0


workbook = openpyxl.load_workbook('/home/yudiaramos/Desktop/projetos/trunks-test/inicial.xlsx')
sheet = workbook.active

# Exemplo de chamada da função para calcular a situação do aluno na linha 4
calcular_situacao_aluno(4)

# Salvar as alterações no arquivo
workbook.save('/home/yudiaramos/Desktop/projetos/trunks-test/saida-final.xlsx') 

